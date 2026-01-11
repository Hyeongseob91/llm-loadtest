"""Benchmark API routes."""

import csv
import io
import json
from datetime import datetime
from typing import Literal, Optional

import httpx
from fastapi import APIRouter, HTTPException, Depends, Query
from fastapi.responses import Response, StreamingResponse

from llm_loadtest_api.auth import APIKeyAuth
from llm_loadtest_api.database import Database
from llm_loadtest_api.logging_config import get_logger, log_benchmark_event
from llm_loadtest_api.services.benchmark_service import BenchmarkService
from llm_loadtest_api.models.schemas import (
    BenchmarkRequest,
    BenchmarkResponse,
    BenchmarkStatus,
    CompareRequest,
    CompareResponse,
    HealthResponse,
    RunListResponse,
)
from llm_loadtest_api import __version__

# Logger
logger = get_logger(__name__)


router = APIRouter(prefix="/api/v1/benchmark", tags=["benchmark"])

# Database and service instances (singleton pattern)
_db: Optional[Database] = None
_service: Optional[BenchmarkService] = None


def get_db() -> Database:
    """Get database instance."""
    global _db
    if _db is None:
        _db = Database()
    return _db


def get_service() -> BenchmarkService:
    """Get benchmark service instance."""
    global _service
    if _service is None:
        _service = BenchmarkService(get_db())
    return _service


@router.get("/health", response_model=HealthResponse)
async def health_check() -> HealthResponse:
    """Check API health."""
    return HealthResponse(
        status="healthy",
        version=__version__,
        timestamp=datetime.now(),
    )


@router.post("/run", response_model=dict, dependencies=[Depends(APIKeyAuth(required=True))])
async def start_benchmark(
    request: BenchmarkRequest,
    service: BenchmarkService = Depends(get_service),
) -> dict:
    """Start a new benchmark run.

    Requires API key authentication via X-API-Key header.

    Returns the run_id for tracking progress.
    """
    logger.info(
        "benchmark_request_received",
        server_url=request.server_url,
        model=request.model,
        adapter=request.adapter,
    )

    run_id = await service.start_benchmark(request)

    log_benchmark_event(
        "benchmark_started",
        run_id=run_id,
        extra={
            "server_url": request.server_url,
            "model": request.model,
            "adapter": request.adapter,
        },
    )

    return {"run_id": run_id, "status": "started"}


@router.get("/run/{run_id}", response_model=BenchmarkStatus)
async def get_run_status(
    run_id: str,
    service: BenchmarkService = Depends(get_service),
) -> BenchmarkStatus:
    """Get benchmark run status."""
    run = service.get_status(run_id)
    if not run:
        raise HTTPException(status_code=404, detail="Run not found")

    return BenchmarkStatus(
        run_id=run["id"],
        status=run["status"],
        server_url=run["server_url"],
        model=run["model"],
        adapter=run["adapter"],
        started_at=run.get("started_at"),
        completed_at=run.get("completed_at"),
        created_at=run["created_at"],
    )


@router.get("/result/{run_id}")
async def get_run_result(
    run_id: str,
    service: BenchmarkService = Depends(get_service),
) -> dict:
    """Get benchmark result."""
    result = service.get_result(run_id)
    if not result:
        # Check if run exists but not completed
        run = service.get_status(run_id)
        if not run:
            raise HTTPException(status_code=404, detail="Run not found")
        if run["status"] == "running":
            raise HTTPException(status_code=202, detail="Benchmark still running")
        if run["status"] == "failed":
            raise HTTPException(status_code=500, detail="Benchmark failed")
        raise HTTPException(status_code=404, detail="Result not found")

    # Add summary if not present
    if "summary" not in result and "results" in result:
        results = result["results"]
        if results:
            best_throughput = max(r["throughput_tokens_per_sec"] for r in results)
            best_ttft = min(r["ttft"]["p50"] for r in results)
            best_result = max(results, key=lambda r: r["throughput_tokens_per_sec"])
            total_requests = sum(r["total_requests"] for r in results)
            failed_requests = sum(r["failed_requests"] for r in results)

            result["summary"] = {
                "best_throughput": best_throughput,
                "best_ttft_p50": best_ttft,
                "best_concurrency": best_result["concurrency"],
                "total_requests": total_requests,
                "overall_error_rate": (failed_requests / total_requests * 100) if total_requests > 0 else 0,
            }

            # Add Goodput if available
            goodput_results = [r["goodput"] for r in results if r.get("goodput")]
            if goodput_results:
                avg_goodput = sum(g["goodput_percent"] for g in goodput_results) / len(goodput_results)
                result["summary"]["avg_goodput_percent"] = avg_goodput

    return result


@router.get("/history", response_model=RunListResponse)
async def list_runs(
    limit: int = 50,
    offset: int = 0,
    status: Optional[str] = None,
    service: BenchmarkService = Depends(get_service),
) -> RunListResponse:
    """List benchmark runs."""
    runs = service.list_runs(limit, offset, status)

    return RunListResponse(
        runs=[
            BenchmarkStatus(
                run_id=r["id"],
                status=r["status"],
                server_url=r["server_url"],
                model=r["model"],
                adapter=r["adapter"],
                started_at=r.get("started_at"),
                completed_at=r.get("completed_at"),
                created_at=r["created_at"],
            )
            for r in runs
        ],
        total=len(runs),  # TODO: Get actual total count
        limit=limit,
        offset=offset,
    )


@router.delete("/run/{run_id}", dependencies=[Depends(APIKeyAuth(required=True))])
async def delete_run(
    run_id: str,
    service: BenchmarkService = Depends(get_service),
) -> dict:
    """Delete a benchmark run.

    Requires API key authentication via X-API-Key header.
    """
    logger.info("benchmark_delete_request", run_id=run_id)

    deleted = service.delete_run(run_id)
    if not deleted:
        raise HTTPException(status_code=404, detail="Run not found")

    log_benchmark_event("benchmark_deleted", run_id=run_id)

    return {"deleted": run_id}


@router.post("/compare")
async def compare_runs(
    request: CompareRequest,
    service: BenchmarkService = Depends(get_service),
) -> dict:
    """Compare multiple benchmark runs."""
    comparison = service.compare_runs(request.run_ids)
    return comparison


@router.get("/result/{run_id}/export")
async def export_result(
    run_id: str,
    format: Literal["csv", "xlsx"] = "csv",
    service: BenchmarkService = Depends(get_service),
) -> Response:
    """Export benchmark result to CSV or Excel format.

    Args:
        run_id: The benchmark run ID.
        format: Export format ('csv' or 'xlsx').

    Returns:
        File download response.
    """
    result = service.get_result(run_id)
    if not result:
        run = service.get_status(run_id)
        if not run:
            raise HTTPException(status_code=404, detail="Run not found")
        if run["status"] == "running":
            raise HTTPException(status_code=202, detail="Benchmark still running")
        if run["status"] == "failed":
            raise HTTPException(status_code=500, detail="Benchmark failed")
        raise HTTPException(status_code=404, detail="Result not found")

    if format == "csv":
        content = _export_to_csv(result)
        return Response(
            content=content,
            media_type="text/csv",
            headers={
                "Content-Disposition": f'attachment; filename="{run_id[:8]}_benchmark.csv"'
            },
        )
    else:
        content = _export_to_xlsx(result)
        return Response(
            content=content,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{run_id[:8]}_benchmark.xlsx"'
            },
        )


def _export_to_csv(result: dict) -> str:
    """Convert benchmark result to CSV format."""
    output = io.StringIO()
    writer = csv.writer(output)

    # Metadata section
    writer.writerow(["# Benchmark Result Export"])
    writer.writerow(["Run ID", result.get("run_id", "N/A")])
    writer.writerow(["Model", result.get("model", "N/A")])
    writer.writerow(["Server URL", result.get("server_url", "N/A")])
    writer.writerow(["Adapter", result.get("adapter", "N/A")])
    writer.writerow(["Started At", result.get("started_at", "N/A")])
    writer.writerow(["Completed At", result.get("completed_at", "N/A")])
    writer.writerow(["Duration (s)", result.get("duration_seconds", "N/A")])
    writer.writerow([])

    # Summary section
    summary = result.get("summary", {})
    writer.writerow(["# Summary"])
    writer.writerow(["Best Throughput (tok/s)", summary.get("best_throughput", "N/A")])
    writer.writerow(["Best TTFT p50 (ms)", summary.get("best_ttft_p50", "N/A")])
    writer.writerow(["Best Concurrency", summary.get("best_concurrency", "N/A")])
    writer.writerow(["Total Requests", summary.get("total_requests", "N/A")])
    writer.writerow(["Overall Error Rate (%)", summary.get("overall_error_rate", "N/A")])
    if summary.get("avg_goodput_percent") is not None:
        writer.writerow(["Avg Goodput (%)", summary.get("avg_goodput_percent")])
    writer.writerow([])

    # Results table
    writer.writerow(["# Detailed Results by Concurrency"])
    writer.writerow([
        "Concurrency",
        "Throughput (tok/s)",
        "Request Rate (req/s)",
        "TTFT p50 (ms)",
        "TTFT p95 (ms)",
        "TTFT p99 (ms)",
        "TPOT p50 (ms)",
        "TPOT p95 (ms)",
        "TPOT p99 (ms)",
        "E2E p50 (ms)",
        "E2E p95 (ms)",
        "E2E p99 (ms)",
        "Total Requests",
        "Successful",
        "Failed",
        "Error Rate (%)",
        "Goodput (%)",
    ])

    for r in result.get("results", []):
        ttft = r.get("ttft", {})
        tpot = r.get("tpot", {})
        e2e = r.get("e2e_latency", {})
        goodput = r.get("goodput", {})

        writer.writerow([
            r.get("concurrency", "N/A"),
            f"{r.get('throughput_tokens_per_sec', 0):.2f}",
            f"{r.get('request_rate_per_sec', 0):.2f}",
            f"{ttft.get('p50', 0):.2f}",
            f"{ttft.get('p95', 0):.2f}",
            f"{ttft.get('p99', 0):.2f}",
            f"{tpot.get('p50', 0):.2f}" if tpot else "N/A",
            f"{tpot.get('p95', 0):.2f}" if tpot else "N/A",
            f"{tpot.get('p99', 0):.2f}" if tpot else "N/A",
            f"{e2e.get('p50', 0):.2f}",
            f"{e2e.get('p95', 0):.2f}",
            f"{e2e.get('p99', 0):.2f}",
            r.get("total_requests", 0),
            r.get("successful_requests", 0),
            r.get("failed_requests", 0),
            f"{r.get('error_rate_percent', 0):.2f}",
            f"{goodput.get('goodput_percent', 0):.2f}" if goodput else "N/A",
        ])

    return output.getvalue()


def _export_to_xlsx(result: dict) -> bytes:
    """Convert benchmark result to Excel format."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="Excel export not available. Install openpyxl: pip install openpyxl",
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Benchmark Results"

    # Styles
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF")

    row = 1

    # Metadata section
    ws.cell(row=row, column=1, value="Benchmark Result Export").font = Font(bold=True, size=14)
    row += 2

    metadata = [
        ("Run ID", result.get("run_id", "N/A")),
        ("Model", result.get("model", "N/A")),
        ("Server URL", result.get("server_url", "N/A")),
        ("Adapter", result.get("adapter", "N/A")),
        ("Started At", result.get("started_at", "N/A")),
        ("Completed At", result.get("completed_at", "N/A")),
        ("Duration (s)", result.get("duration_seconds", "N/A")),
    ]

    for label, value in metadata:
        ws.cell(row=row, column=1, value=label).font = header_font
        ws.cell(row=row, column=2, value=value)
        row += 1

    row += 1

    # Summary section
    ws.cell(row=row, column=1, value="Summary").font = Font(bold=True, size=12)
    row += 1

    summary = result.get("summary", {})
    summary_data = [
        ("Best Throughput (tok/s)", summary.get("best_throughput", "N/A")),
        ("Best TTFT p50 (ms)", summary.get("best_ttft_p50", "N/A")),
        ("Best Concurrency", summary.get("best_concurrency", "N/A")),
        ("Total Requests", summary.get("total_requests", "N/A")),
        ("Overall Error Rate (%)", summary.get("overall_error_rate", "N/A")),
    ]
    if summary.get("avg_goodput_percent") is not None:
        summary_data.append(("Avg Goodput (%)", summary.get("avg_goodput_percent")))

    for label, value in summary_data:
        ws.cell(row=row, column=1, value=label).font = header_font
        ws.cell(row=row, column=2, value=value)
        row += 1

    row += 2

    # Results table
    ws.cell(row=row, column=1, value="Detailed Results by Concurrency").font = Font(bold=True, size=12)
    row += 1

    headers = [
        "Concurrency", "Throughput", "Req Rate", "TTFT p50", "TTFT p95", "TTFT p99",
        "TPOT p50", "TPOT p95", "TPOT p99", "E2E p50", "E2E p95", "E2E p99",
        "Total", "Success", "Failed", "Error %", "Goodput %",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    row += 1

    for r in result.get("results", []):
        ttft = r.get("ttft", {})
        tpot = r.get("tpot", {})
        e2e = r.get("e2e_latency", {})
        goodput = r.get("goodput", {})

        values = [
            r.get("concurrency", 0),
            round(r.get("throughput_tokens_per_sec", 0), 2),
            round(r.get("request_rate_per_sec", 0), 2),
            round(ttft.get("p50", 0), 2),
            round(ttft.get("p95", 0), 2),
            round(ttft.get("p99", 0), 2),
            round(tpot.get("p50", 0), 2) if tpot else None,
            round(tpot.get("p95", 0), 2) if tpot else None,
            round(tpot.get("p99", 0), 2) if tpot else None,
            round(e2e.get("p50", 0), 2),
            round(e2e.get("p95", 0), 2),
            round(e2e.get("p99", 0), 2),
            r.get("total_requests", 0),
            r.get("successful_requests", 0),
            r.get("failed_requests", 0),
            round(r.get("error_rate_percent", 0), 2),
            round(goodput.get("goodput_percent", 0), 2) if goodput else None,
        ]

        for col, value in enumerate(values, 1):
            ws.cell(row=row, column=col, value=value)
        row += 1

    # Adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 12

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


@router.get("/result/{run_id}/analysis")
async def analyze_result(
    run_id: str,
    server_url: str = Query(default="http://host.docker.internal:8000", description="vLLM server URL"),
    model: str = Query(default="", description="Model name (uses benchmark model if empty)"),
    service: BenchmarkService = Depends(get_service),
) -> StreamingResponse:
    """Generate AI analysis of benchmark results using vLLM.

    Streams the analysis response in real-time using Server-Sent Events (SSE).

    Args:
        run_id: The benchmark run ID.
        server_url: vLLM server URL for analysis generation.
        model: Model to use for analysis (defaults to benchmark's model).

    Returns:
        StreamingResponse with SSE format.
    """
    result = service.get_result(run_id)
    if not result:
        run = service.get_status(run_id)
        if not run:
            raise HTTPException(status_code=404, detail="Run not found")
        if run["status"] == "running":
            raise HTTPException(status_code=202, detail="Benchmark still running")
        if run["status"] == "failed":
            raise HTTPException(status_code=500, detail="Benchmark failed")
        raise HTTPException(status_code=404, detail="Result not found")

    # Use model from result if not specified
    analysis_model = model if model else result.get("model", "qwen3-14b")

    # Build analysis prompt
    prompt = _build_analysis_prompt(result)

    async def generate_analysis():
        """Stream analysis from vLLM."""
        try:
            async with httpx.AsyncClient(timeout=httpx.Timeout(300.0, read=300.0)) as client:
                async with client.stream(
                    "POST",
                    f"{server_url}/v1/chat/completions",
                    json={
                        "model": analysis_model,
                        "messages": [
                            {
                                "role": "system",
                                "content": "당신은 LLM 서버 성능 분석 전문가입니다. 벤치마크 결과를 분석하여 마크다운 형식의 보고서를 작성합니다. 한국어로 답변하세요."
                            },
                            {
                                "role": "user",
                                "content": prompt
                            }
                        ],
                        "stream": True,
                        "max_tokens": 8192,
                        "temperature": 0.3,
                    },
                    headers={"Content-Type": "application/json"},
                ) as response:
                    if response.status_code != 200:
                        error_text = await response.aread()
                        yield f"data: {json.dumps({'error': f'vLLM error: {error_text.decode()}'})}\n\n"
                        return

                    # Thinking 모델 대응: 실제 보고서 시작 전까지 버퍼링
                    buffer = ""
                    report_started = False

                    async for line in response.aiter_lines():
                        if line.startswith("data: "):
                            data = line[6:]
                            if data == "[DONE]":
                                # 버퍼에 남은 내용이 있으면 출력
                                if buffer and not report_started:
                                    yield f"data: {json.dumps({'content': buffer})}\n\n"
                                yield "data: [DONE]\n\n"
                                break
                            try:
                                chunk = json.loads(data)
                                content = chunk.get("choices", [{}])[0].get("delta", {}).get("content", "")
                                if content:
                                    if report_started:
                                        # 보고서 시작된 후에는 바로 출력
                                        yield f"data: {json.dumps({'content': content})}\n\n"
                                    else:
                                        # 버퍼에 축적
                                        buffer += content
                                        # 마크다운 헤딩이 나오면 보고서 시작
                                        if "\n#" in buffer or buffer.startswith("#"):
                                            # # 이전의 thinking 부분 제거
                                            idx = buffer.find("\n#")
                                            if idx != -1:
                                                buffer = buffer[idx + 1:]  # \n 제거하고 # 부터
                                            report_started = True
                                            yield f"data: {json.dumps({'content': buffer})}\n\n"
                                            buffer = ""
                            except json.JSONDecodeError:
                                continue
        except httpx.ConnectError:
            yield f"data: {json.dumps({'error': f'vLLM 서버에 연결할 수 없습니다: {server_url}'})}\n\n"
        except Exception as e:
            yield f"data: {json.dumps({'error': str(e)})}\n\n"

    return StreamingResponse(
        generate_analysis(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
        },
    )


def _build_analysis_prompt(result: dict) -> str:
    """Build analysis prompt from benchmark result."""
    model = result.get("model", "Unknown")
    server_url = result.get("server_url", "Unknown")
    duration = result.get("duration_seconds", 0)
    results = result.get("results", [])
    summary = result.get("summary", {})

    # Build concurrency results table
    results_table = "| Concurrency | Throughput (tok/s) | TTFT p50 (ms) | TTFT p99 (ms) | Error Rate (%) | Goodput (%) |\n"
    results_table += "|-------------|-------------------|---------------|---------------|----------------|-------------|\n"

    for r in results:
        ttft = r.get("ttft") or {}
        goodput = r.get("goodput") or {}
        goodput_str = f"{goodput.get('goodput_percent', 0):.1f}" if goodput else "N/A"
        results_table += f"| {r.get('concurrency', 0)} | {r.get('throughput_tokens_per_sec', 0):.1f} | {ttft.get('p50', 0):.1f} | {ttft.get('p99', 0):.1f} | {r.get('error_rate_percent', 0):.2f} | {goodput_str} |\n"

    prompt = f"""다음 LLM 서버 벤치마크 결과를 분석해주세요.

## 테스트 환경
- **모델**: {model}
- **서버**: {server_url}
- **테스트 시간**: {duration:.1f}초

## 결과 요약
- **최고 처리량**: {summary.get('best_throughput', 0):.1f} tok/s
- **최저 TTFT (p50)**: {summary.get('best_ttft_p50', 0):.1f} ms
- **최적 동시성**: {summary.get('best_concurrency', 'N/A')}
- **전체 에러율**: {summary.get('overall_error_rate', 0):.2f}%
- **평균 Goodput**: {summary.get('avg_goodput_percent', 'N/A')}%

## Concurrency별 상세 결과
{results_table}

## 분석 요청
위 벤치마크 결과를 바탕으로 다음 항목들을 분석해주세요:

1. **성능 개요**: 전반적인 서버 성능 평가
2. **Concurrency 영향 분석**: 동시성 증가에 따른 성능 변화 패턴
3. **병목 지점 식별**: 성능이 저하되기 시작하는 동시성 레벨과 원인 추정
4. **TTFT vs Throughput 트레이드오프**: 응답 시간과 처리량 간의 관계 분석
5. **권장 운영 동시성**: 실제 서비스에서 권장하는 최적 동시성 레벨
6. **개선 제안**: 성능 향상을 위한 구체적인 제안

각 항목을 **마크다운 형식**으로 작성해주세요."""

    return prompt
